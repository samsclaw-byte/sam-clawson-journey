#!/usr/bin/env python3
"""
Generate Work Task Management Excel Template
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import calendar

def create_work_template():
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styles
    header_fill = PatternFill(start_color="9945FF", end_color="9945FF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="E8D5F5", end_color="E8D5F5", fill_type="solid")
    subheader_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    # === SHEET 1: Task Input ===
    ws1 = wb.create_sheet("Task Input")
    
    # Headers
    task_headers = [
        "Item Name", "Type", "Detail", "TAT Days", "Due Date", 
        "Category", "Source", "Project", "Stakeholder", 
        "Assigned To", "Status", "Priority", "Effort", "Blocked Reason"
    ]
    
    for col, header in enumerate(task_headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Sample data rows
    sample_tasks = [
        ["MGC fees for digital card", "Task", "Create unit comparison DGC vs Physical, Scenario table assuming avg card load, variable price scenario table and incrementality assuming 2025 volumes", 3, "=TODAY()+3", "Analysis", "In Person", "Digital Card Strategy", "Dem", "Me", "In Progress", "P1", "Medium", "N/A"],
        ["Aylin share card campaign invoicing", "Task", "Process invoice for card campaign", 3, "=TODAY()+3", "Financial", "Email", "Card Campaign", "Aylin", "Me", "In Progress", "P1", "Small", "N/A"],
        ["Complete my success tasks", "Task", "Finish assigned success tasks", 3, "=TODAY()+3", "Admin", "Email", "N/A", "Colin", "Me", "Not Started", "P2", "Medium", "N/A"],
        ["No objection certificate", "Task", "Obtain NOC for project approval", 7, "=TODAY()+7", "Admin", "Email", "Project X", "Ahmed", "Me", "Not Started", "P1", "Small", "N/A"],
        ["Umair contract renewal", "Project", "Contract renewal process with Umair", 7, "=TODAY()+7", "Operations", "In Person", "Osman Contract", "Osman", "Me", "In Progress", "P0", "Large", "Awaiting legal review"],
        ["Raffy fully loaded marketing spend", "Task", "Calculate fully loaded marketing spend", 7, "=TODAY()+7", "Financial", "In Person", "Marketing Q1", "Raffy", "Me", "Complete", "P1", "Medium", "N/A"],
    ]
    
    for row_idx, task in enumerate(sample_tasks, 2):
        for col_idx, value in enumerate(task, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            
            # Format Due Date column
            if col_idx == 5:  # Due Date
                cell.number_format = 'YYYY-MM-DD'
    
    # Column widths
    col_widths = [30, 12, 50, 12, 12, 15, 12, 25, 15, 15, 15, 12, 12, 30]
    for idx, width in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(idx)].width = width
    
    # Add data validation
    # Type dropdown
    type_validation = DataValidation(type="list", formula1='"Task,Project,Recurring"', allow_blank=False)
    type_validation.add(f'B2:B100')
    ws1.add_data_validation(type_validation)
    
    # Category dropdown
    cat_validation = DataValidation(type="list", formula1='"Analysis,Financial,Admin,Campaigns,Operations,Strategy"', allow_blank=True)
    cat_validation.add(f'F2:F100')
    ws1.add_data_validation(cat_validation)
    
    # Source dropdown
    source_validation = DataValidation(type="list", formula1='"Email,In Person,Teams,Phone,System,Other"', allow_blank=True)
    source_validation.add(f'G2:G100')
    ws1.add_data_validation(source_validation)
    
    # Status dropdown
    status_validation = DataValidation(type="list", formula1='"Not Started,In Progress,Blocked,Complete"', allow_blank=False)
    status_validation.add(f'K2:K100')
    ws1.add_data_validation(status_validation)
    
    # Priority dropdown
    priority_validation = DataValidation(type="list", formula1='"P0,P1,P2,P3"', allow_blank=False)
    priority_validation.add(f'L2:L100')
    ws1.add_data_validation(priority_validation)
    
    # Effort dropdown
    effort_validation = DataValidation(type="list", formula1='"Small,Medium,Large"', allow_blank=False)
    effort_validation.add(f'M2:M100')
    ws1.add_data_validation(effort_validation)
    
    # Add note
    ws1['A15'] = "INSTRUCTIONS: Fill in rows above. Copy the entire row and paste into Telegram chat. Due Date auto-calculates from TAT Days."
    ws1['A15'].font = Font(italic=True, color="666666", size=9)
    
    # Freeze header row
    ws1.freeze_panes = 'A2'
    
    # === SHEET 2: Recurring Items ===
    ws2 = wb.create_sheet("Recurring Items")
    
    recurring_headers = [
        "Item Name", "Type", "Frequency", "Category", "Assigned To", 
        "Stakeholder", "Template/Checklist", "Next Due", "Status"
    ]
    
    for col, header in enumerate(recurring_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Sample recurring items
    recurring_data = [
        ["Month-end finance close", "Recurring", "Monthly", "Financial", "Me", "CFO", "1. GL entries 2. Reconciliations 3. Variance analysis 4. Management pack", "2026-02-28", "Not Started"],
        ["Marketing spend report", "Recurring", "Monthly", "Financial", "Me", "Raffy", "1. Collect channel data 2. Calculate ROI 3. Compare to budget 4. Flag anomalies", "2026-03-05", "Not Started"],
        ["Team performance review", "Recurring", "Quarterly", "Operations", "Me", "Manager", "1. Review KPIs 2. 1:1s with team 3. Documentation 4. Action items", "2026-03-31", "Not Started"],
    ]
    
    for row_idx, item in enumerate(recurring_data, 2):
        for col_idx, value in enumerate(item, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    # Column widths for recurring
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 15
    ws2.column_dimensions['E'].width = 15
    ws2.column_dimensions['F'].width = 15
    ws2.column_dimensions['G'].width = 50
    ws2.column_dimensions['H'].width = 12
    ws2.column_dimensions['I'].width = 15
    
    # Add validation
    freq_validation = DataValidation(type="list", formula1='"Daily,Weekly,Monthly,Quarterly,Annually"', allow_blank=False)
    freq_validation.add(f'C2:C100')
    ws2.add_data_validation(freq_validation)
    
    status_validation2 = DataValidation(type="list", formula1='"Not Started,In Progress,Complete"', allow_blank=False)
    status_validation2.add(f'I2:I100')
    ws2.add_data_validation(status_validation2)
    
    ws2.freeze_panes = 'A2'
    
    # === SHEET 3: Project Master List ===
    ws3 = wb.create_sheet("Project Master")
    
    project_headers = [
        "Project Name", "Description", "Status", "Start Date", 
        "Target End", "Progress %", "Lead", "Stakeholders"
    ]
    
    for col, header in enumerate(project_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Sample projects
    project_data = [
        ["Digital Card Strategy", "End-to-end digital card product launch strategy", "In Progress", "2026-02-01", "2026-03-15", 45, "Me", "Dem, Aylin, Raffy"],
        ["Osman Contract Renewal", "Contract renewal and renegotiation", "In Progress", "2026-02-10", "2026-02-28", 30, "Me", "Osman, Legal"],
        ["Card Campaign", "Marketing campaign for card products", "In Progress", "2026-02-01", "2026-02-20", 70, "Aylin", "Me, Dem"],
        ["Marketing Q1", "Q1 marketing initiatives and spend tracking", "In Progress", "2026-01-01", "2026-03-31", 60, "Raffy", "Me, Dem"],
    ]
    
    for row_idx, proj in enumerate(project_data, 2):
        for col_idx, value in enumerate(proj, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    # Column widths
    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 40
    ws3.column_dimensions['C'].width = 15
    ws3.column_dimensions['D'].width = 12
    ws3.column_dimensions['E'].width = 12
    ws3.column_dimensions['F'].width = 12
    ws3.column_dimensions['G'].width = 15
    ws3.column_dimensions['H'].width = 30
    
    # Validation
    proj_status_validation = DataValidation(type="list", formula1='"Planning,In Progress,On Hold,Complete"', allow_blank=False)
    proj_status_validation.add(f'C2:C100')
    ws3.add_data_validation(proj_status_validation)
    
    ws3.freeze_panes = 'A2'
    
    # === SHEET 4: Recurring Schedule View ===
    ws4 = wb.create_sheet("Recurring Schedule")
    
    ws4['A1'] = "RECURRING ITEMS SCHEDULE VIEW"
    ws4['A1'].font = Font(bold=True, size=14, color="9945FF")
    ws4.merge_cells('A1:F1')
    
    ws4['A3'] = "February 2026"
    ws4['A3'].font = Font(bold=True, size=12)
    
    # Calendar headers
    days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    for col, day in enumerate(days, 1):
        cell = ws4.cell(row=4, column=col, value=day)
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # February 2026 calendar (simplified)
    # Feb 1 = Sunday
    feb_dates = [
        ["", "", "", "", "", "", "1"],
        ["2", "3", "4", "5", "6", "7", "8"],
        ["9", "10", "11", "12", "13", "14", "15"],
        ["16", "17", "18", "19", "20", "21", "22"],
        ["23", "24", "25", "26", "27", "28", ""],
    ]
    
    for row_idx, week in enumerate(feb_dates, 5):
        for col_idx, date in enumerate(week, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=date)
            cell.alignment = Alignment(horizontal='center', vertical='top')
            cell.border = border
            if date == "28":
                cell.fill = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
                cell.value = "28\nMonth-end\nclose"
                cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    
    # Set row heights for calendar
    for row in range(5, 10):
        ws4.row_dimensions[row].height = 40
    
    for col in range(1, 8):
        ws4.column_dimensions[get_column_letter(col)].width = 15
    
    # Upcoming recurring list
    ws4['A12'] = "UPCOMING RECURRING ITEMS"
    ws4['A12'].font = Font(bold=True, size=12)
    
    upcoming_headers = ["Date", "Item", "Frequency", "Assigned"]
    for col, header in enumerate(upcoming_headers, 1):
        cell = ws4.cell(row=13, column=col, value=header)
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.border = border
    
    upcoming_data = [
        ["Feb 28", "Month-end finance close", "Monthly", "Me"],
        ["Mar 5", "Marketing spend report", "Monthly", "Me"],
        ["Mar 31", "Team performance review", "Quarterly", "Me"],
        ["Mar 31", "Month-end finance close", "Monthly", "Me"],
    ]
    
    for row_idx, item in enumerate(upcoming_data, 14):
        for col_idx, value in enumerate(item, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
    
    for col in [1, 2, 3, 4]:
        ws4.column_dimensions[get_column_letter(col)].width = 20 if col == 2 else 15
    
    # Save workbook
    output_path = '/home/samsclaw/.openclaw/workspace/work-tasks-template.xlsx'
    wb.save(output_path)
    print(f"✅ Excel template created: {output_path}")
    print(f"\n📊 Sheets created:")
    print("   1. Task Input - Main entry sheet with validation")
    print("   2. Recurring Items - Recurring deliverables")
    print("   3. Project Master - Project overview list")
    print("   4. Recurring Schedule - Calendar view of upcoming items")
    print(f"\n🎯 Features:")
    print("   • Dropdown validation for Type, Category, Status, Priority, Effort")
    print("   • Auto-calculated Due Date from TAT Days")
    print("   • Sample data in each sheet")
    print("   • Copy-paste ready for Telegram")

if __name__ == '__main__':
    create_work_template()
